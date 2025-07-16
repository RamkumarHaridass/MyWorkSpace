import pandas as pd
import argparse
import logging
import openpyxl
import xlsxwriter
from openpyxl.styles import PatternFill
import reportCreator

def convertxlsToxlsx(file):
    logging.info("Converting Excel from .xls to .xlsx")
    xlsFile = pd.read_excel(file, sheet_name=None)
    xlsxFile = file + 'x'
    writer = pd.ExcelWriter(xlsxFile, engine='openpyxl')
    for sheet_name, df in xlsFile.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.close()

def convertxlsmToxlsx(file):
    logging.info("Converting Excel from .xlsm to .xlsx")
    xlsmFile = pd.read_excel(file, sheet_name=None)
    xlsxFile = file.replace('.xlsm', '.xlsx')
    writer = pd.ExcelWriter(xlsxFile, engine='openpyxl')
    for sheet_name, df in xlsmFile.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.close()

def excelCompareOpenPyxl(file1, file2, roundOffValue):
    numberOfDiffsSheetLevel = {}
    changedList = []
    changesDictSheetLevel = {}
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)
    sheets1 = set(wb1.sheetnames)
    sheets2 = set(wb2.sheetnames)
    if sheets1 - sheets1:
        logging.info(f"Sheet names {sheets1 - sheets2} not found in {file2}")
    if sheets2 - sheets1:
        logging.info(f"Sheet names {sheets2 - sheets1} not found in {file1}")
    commonSheets = sheets1.intersection(sheets2)
    if len(commonSheets) == 0:
        return -1, 'No Common sheets found', changesDictSheetLevel, numberOfDiffsSheetLevel
    for sheet in commonSheets:
        logging.info(f"Comparing sheet start {sheet}")
        changesDataLevel = {}
        ws1 = wb1[sheet]
        ws2 = wb2[sheet]
        max_row = max(ws1.max_row, ws2.max_row)
        max_col = max(ws1.max_column, ws2.max_column)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell1 = ws1.cell(row=row, column=col).value
                cell2 = ws2.cell(row=row, column=col).value
                try:
                    if roundOffValue != 'NA':
                        if round(float(cell1), int(roundOffValue)) != round(float(cell2), int(roundOffValue)):
                            columnExcelName = xlsxwriter.utility.xl_col_to_name(col-1)
                            changesDataLevel[(columnExcelName+str(row))] = (cell1, cell2)
                            changedList.append((sheet, row, col, cell1, cell2))
                    else:
                        if cell1 != cell2:
                            columnExcelName = xlsxwriter.utility.xl_col_to_name(col - 1)
                            changesDataLevel[(columnExcelName + str(row))] = (cell1, cell2)
                            changedList.append((sheet, row, col, cell1, cell2))
                except Exception:
                    if cell1 != cell2:
                        columnExcelName = xlsxwriter.utility.xl_col_to_name(col-1)
                        changesDataLevel[(columnExcelName+str(row))] = (cell1, cell2)
                        changedList.append((sheet, row, col, cell1, cell2))
        changesDictSheetLevel[sheet] = changesDataLevel
        if changesDataLevel:
            numberOfDiffsSheetLevel[sheet] = len(changesDataLevel)
        logging.info(f"Comparing sheet end {sheet}")
    wb1.close()
    wb2.close()
    return len(changedList), 'compare complete', changesDictSheetLevel, numberOfDiffsSheetLevel

def reportGenerationHtml(htmlFile, differedDict):
    style = reportCreator.reportStyle()
    header = f""" </head>
        <body>
        <table id=\"reportTable\" class=\"table\">
                    <thead>
                        <tr>
                            <th><button class=\"table__header\">Sheet Name</button></th>
                            <th><button class=\"table__header\">Differed Cell Name</button></th>
                            <th><button class=\"table__header\">Base Value</button></th>    
                            <th><button class=\"table__header\">Actual Value</button></th>                                                                               
                        </tr>
                    </thead>"""
    htmlContent = style + header
    with open(htmlFile, "w", encoding='utf-8') as hFile:
        hFile.write(htmlContent)
        for sheet in differedDict:
            for cell in differedDict[sheet]:
                baseValue, actualValue = differedDict[sheet][cell]
                hFile.write(f"""<tr>
                            <td>{sheet}</td>
                            <td>{cell}</td>
                            <td>{baseValue}</td>
                            <td>{actualValue}</td>
                        </tr>""")
        hFile.write("</table></body></html>")

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-b', '--baseFileLocation', help='Base file location with directory')
    parser.add_argument('-a', '--actualFileLocation', help='Actual file location with directory')
    parser.add_argument('-o', '--outputFileLocation', help='Output file location with directory')
    parser.add_argument('-r', '--roundOffValue', help='Number of digits to round off before compare', default='NA')
    args = parser.parse_args()
    logging.basicConfig(format='%(message)s', level=logging.INFO)
    baseFile = args.baseFileLocation
    actualFile = args.actualFileLocation
    outFile = args.outputFileLocation
    roundOffValue = args.roundOffValue

    logging.info('***** Excel Comparison started *****')

    if baseFile.endswith('.xls'):
        convertxlsToxlsx(baseFile)
        baseFile += 'x'
    if actualFile.endswith('.xls'):
        convertxlsToxlsx(actualFile)
        actualFile += 'x'
    if baseFile.endswith('.xlsm'):
        convertxlsmToxlsx(baseFile)
        baseFile = baseFile.replace('.xlsm', '.xlsx')
    if actualFile.endswith('.xlsm'):
        convertxlsmToxlsx(actualFile)
        actualFile = actualFile.replace('.xlsm', '.xlsx')

    differedValues, message, differedDict, numberOfDiffsSheetLevel = excelCompareOpenPyxl(baseFile, actualFile, roundOffValue)

    if numberOfDiffsSheetLevel:
        logging.info(f"***** Differences observed in {len(numberOfDiffsSheetLevel)} sheets : {numberOfDiffsSheetLevel} *****")
    if differedValues == 0:
        logging.info("No differences found in any sheet")
    elif differedValues < 0:
        logging.info(message)
    else:
        reportGenerationHtml(outFile, differedDict)
        logging.info(f"{differedValues} Difference(s) observed")

    logging.info('***** Excel Comparison completed *****')
